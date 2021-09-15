#coding: utf-8

import io, re, datetime
from collections import defaultdict
import requests
import openpyxl
import bokeh.plotting
import bokeh.palettes
bokeh.plotting.output_notebook()


class AmbulanceStats:
    url = 'https://www.fdma.go.jp/disaster/coronavirus/items/coronavirus_data.xlsx'
    pref_column = 2
    area_column = 3
    first_column = 4
    first_row = 6
    week_row = 4

    def __init__(self):
        r = requests.get(self.url)
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        sheet = wb['搬送困難事案（今回）']
        self.data = self._read_sheet(sheet, 2020)
        sheet = wb['搬送困難事案（前年同期）']
        data_prev = self._read_sheet(sheet, 2019)
        self._merge(self.data, data_prev)

    def _read_weeks(self, sheet, first_year):
        # 列と週の対応関係を読む
        column_week = {}
        y = first_year
        column = self.first_column
        week = sheet.cell(column=column, row=self.week_row).value
        while week is not None:
            m1, d1, m2, d2, *_ = [int(x) for x in re.findall('\d+', week)]
            if m1 == 1 and d1 == 1:
                y += 1
            date_start = datetime.date(y, m1, d1)
            if m1 == 12 and m2 == 1:
                y += 1
            date_end = datetime.date(y, m2, d2)
            column_week[column] = {
                'start': date_start,
                'end': date_end,
            }
            column += 1
            week = sheet.cell(column=column, row=self.week_row).value
        return column_week  # column_week[column_number] = {'start': start_date, 'end': end_date}

    def _read_sheet(self, sheet, first_year):
        column_week = self._read_weeks(sheet, first_year)
        
        # データを読む
        data = defaultdict(dict)
        row = self.first_row
        pref = sheet.cell(column=self.pref_column, row=row).value
        area = sheet.cell(column=self.area_column, row=row).value
        while area is not None:
            if area == '52本部合計':
                row += 1
                pref = sheet.cell(column=self.pref_column, row=row).value
                area = sheet.cell(column=self.area_column, row=row).value
                continue
            column = self.first_column
            count = sheet.cell(column=column, row=row).value
            while count is not None:
                week = column_week[column]
                date = week['start']
                dt = datetime.timedelta(days=1)
                while date <= week['end']:
                    data[area][date] = count
                    date += dt
                column += 1
                count = sheet.cell(column=column, row=row).value
            row += 1
            pref = sheet.cell(column=self.pref_column, row=row).value
            area = sheet.cell(column=self.area_column, row=row).value
        return data  # data[area] = {date: count}

    def _merge(self, data1, data2):
        assert data1.keys() == data2.keys()
        for area in data1:
            for date in data2[area]:
                if date not in data1[area]:
                    data1[area][date] = data2[area][date]

    @property
    def areas(self):
        return list(self.data.keys())

    def show(self, area):
        data_area = self.data[area]
        years = sorted({date.year for date in data_area})
        counts = {}
        for year in years:
            date = datetime.date(year, 1, 1)
            counts[year] = [data_area.get(date + datetime.timedelta(days=i), None) for i in range(365)]

        fig = bokeh.plotting.figure(plot_width=800, plot_height=300)
        fig.xaxis.axis_label = '1月1日からの日数'
        fig.yaxis.axis_label = '搬送困難事案の件数'
        for c, year in enumerate(counts):
            count = counts[year]
            x = [i+1 for i in range(len(count)) if count[i] is not None]
            y = [x for x in count if x is not None]
            fig.line(x, y, line_color=bokeh.palettes.Category10[3][c],
              line_alpha=0.7, line_width=3, legend_label=str(year))
        bokeh.plotting.show(fig)
