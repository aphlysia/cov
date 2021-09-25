# coding: utf-8

import re, pathlib, datetime, shelve
import urllib.parse
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import pandas_bokeh
pandas_bokeh.output_notebook()


class PatientStats:
    base = 'https://www.mhlw.go.jp'
    top_url = 'https://www.mhlw.go.jp/stf/seisakunitsuite/newpage_00023.html'
    data_row_start = 8
    data_row_end = 54
    pref_column = 'C'

    def __init__(self, data_dir='data/10900000'):
        self.num_table = str.maketrans('０１２３４５６７８９', '0123456789')
        self.data_dir = pathlib.Path(data_dir)
        self.data_dir.mkdir(exist_ok=True)

    def update(self):
        # ファイル一覧を取得
        r = requests.get(self.top_url)
        assert r.status_code == 200
        soup = BeautifulSoup(r.content, 'html.parser')
        items = [
            (
                self._parse_datetime(item[0]),
                urllib.parse.urljoin(self.base, item[1]['href']),
            )
            for item in zip(
                soup.find(attrs={'class': 'm-grid__col1'}).find_all(text=re.compile(
                    r'新型コロナウイルス感染症患者の療養状況等及び入院患者受入病床数等に関する調査結果.*時点.*')),
                soup.find(attrs={'class': 'm-grid__col1'}).find_all(href=re.compile(r'.xlsx$')),
            )
        ]

        self._download(items)

    def show(self, pref, figsize=(800, 300)):
        org = pd.get_option('plotting.backend')
        pd.set_option('plotting.backend', 'pandas_bokeh')
        self.show_bed_usage(pref, figsize)
        self.show_bed_count(pref, figsize)
        self.show_admission_proportion(pref, figsize)
        self.show_type_count_1(pref, figsize)
        self.show_type_count_2(pref, figsize)
        pd.set_option('plotting.backend', org)

    def _parse_datetime(self, title):
        title = title.translate(self.num_table)  # 全角の数字を半角に
        time_str = re.search(r'\d+年\d+月\d+日\d時', title).group(0)  # 年月日時の部分を切り出す
        return time_str

    def _parse_filename(self, url):
        return urllib.parse.urlparse(url).path.split('/')[-1]

    def _parse_pref(self, text):
        return text.value[3:].replace(' ', '')

    def _download_(self, url, overwrite=False):
        if overwrite == False and (self.data_dir / self._parse_filename(url)).exists():
            return
        print(f'Download {url}')
        r = requests.get(url)
        if r.status_code != 200:
            raise RuntimeError(f'Download failed.\nURL: {url}\nStatus code: {r.status_code}')
        with open(self.data_dir / self._parse_filename(url), 'wb') as f:
            f.write(r.content)

    def _download(self, items):
        with shelve.open(str(self.data_dir / 'files')) as files:
            for dt, url in items:
                # dt は日時, url は xlsx の url
                self._download_(url)
                files[dt] = url

    def _get_header(self, sheet, column, header_row):
        ranges = [r for r in sheet.merged_cells.ranges if f'{column}{header_row}' in r]
        if len(ranges) == 0:
            return sheet[f'{column}{header_row}'].value
        elif len(ranges) == 1:
            for c, r in ranges[0].cells:
                v = sheet.cell(c, r).value
                if v is not None:
                    return v
        return None

    def _validate_header(self, sheet, time, coef, column, header_row, format_date1):
        header = self._get_header(sheet, get_column_letter(column), header_row)
        if coef == '(重症者用でない) 確保病床使用率':
            assert (
                (time >= format_date1 and header == '確保病床\n使用率\n（注５）')
                or (time < format_date1 and header == '確保病床数に対する使用率'))
        elif coef == '(重症者用でない) 確保病床数':
            assert header.startswith('確保病床数\n')
        elif coef == '(重症者用) 確保病床使用率':
            assert (
                (time >= format_date1 and header == '確保病床\n使用率\n（注５）')
                or (time < format_date1 and header == '確保病床数に対する使用率'))
        elif coef == '(重症者用) 確保病床数':
            assert header.startswith('確保病床数\n')
        elif coef == '(宿泊療養施設) 確保居室使用率':
            assert (
                (time >= format_date1 and header == '確保居室\n使用率\n（注９）')
                or (time < format_date1 and header == '確保居室数に対する使用率'))
        elif coef == '(宿泊療養施設) 確保居室数':
            assert header.startswith('確保居室数\n')
        elif coef == '療養者数':
            assert (
                (time >= format_date1 and header == '（１）療養\n者数\n（注１）')
                or (time < format_date1 and header.endswith('（１）PCR検査陽性者数（退院者等除く。）（注１,２）')))
        elif coef == '入院者数':
            assert (
                (time >= format_date1 and header == '（２）①-1\n入院者数')
                or (time < format_date1 and header.endswith('（２）入院者数（入院確定者数を含む）')))
        elif coef == '宿泊療養者数':
            assert (
                (time >= format_date1 and header == '（３）宿泊\n療養者数')
                or (time < format_date1 and header.endswith('（３）宿泊療養者数')))
        elif coef == '自宅療養者等数':
            assert (
                (time >= format_date1 and header == '（４）①-1\n自宅療養者\n等数')
                or (time < format_date1 and header.endswith('\n\n（４）自宅療養者数')))
        elif coef == '療養先調整中の人数':
            assert header == '（５）①-1\n療養先調整\n中の人数\n（注10）'
        elif coef == '入院先調整中の人数':
            assert header == '（５）①-2\nうち、入院\n先調整中の\n人数\n(注11)'

    def _get_stats(self, sheet, row, coef, time):
        header_row = 7
        format_date1 = datetime.datetime(2021, 6, 2, 0, 0)
        # 各指標の列を定義 (000831147.xlsx における列)
        if time >= format_date1:
            stats_column = {
                '(重症者用でない) 確保病床使用率': 10,
                '(重症者用でない) 確保病床数': 9,
                '(重症者用) 確保病床使用率': 17,
                '(重症者用) 確保病床数': 16,
                '(宿泊療養施設) 確保居室使用率': 22,
                '(宿泊療養施設) 確保居室数': 21,
                '療養者数': 4,
                '入院者数': 5,
                '宿泊療養者数': 18,
                '自宅療養者等数': 23,
                '療養先調整中の人数': 25,
                '入院先調整中の人数': 26,
            }
        else:
            stats_column = {
                '(重症者用でない) 確保病床使用率': 8,
                '(重症者用でない) 確保病床数': 7,
                '(重症者用) 確保病床使用率': 13,
                '(重症者用) 確保病床数': 12,
                '(宿泊療養施設) 確保居室使用率': 18,
                '(宿泊療養施設) 確保居室数': 17,
                '療養者数': 4,
                '入院者数': 5,
                '宿泊療養者数': 15,
                '自宅療養者数': 20,
                '社会福士施設等療養者数': 21,
                '療養先調整中の人数': None,
                '入院先調整中の人数': None,
            }
        
        if coef == '入院率':
            a = self._get_stats(sheet, row, '入院者数', time)
            b = self._get_stats(sheet, row, '療養者数', time)
            if not isinstance(a, int) or not isinstance(b, int):
                return None
            if b == 0:
                return 0
            return a / b
        if coef == '宿泊療養である割合':
            a = self._get_stats(sheet, row, '宿泊療養者数', time)
            b = self._get_stats(sheet, row, '療養者数', time)
            if not isinstance(a, int) or not isinstance(b, int):
                return None
            if b == 0:
                return 0
            return a / b
        if coef == '自宅療養である割合':
            a = self._get_stats(sheet, row, '自宅療養者等数', time)
            b = self._get_stats(sheet, row, '療養者数', time)
            if not isinstance(a, int) or not isinstance(b, int):
                return None
            if b == 0:
                return 0
            return a / b
        if coef == '療養先調整中である割合':
            a = self._get_stats(sheet, row, '療養先調整中の人数', time)
            b = self._get_stats(sheet, row, '療養者数', time)
            if not isinstance(a, int) or not isinstance(b, int):
                return None
            if b == 0:
                return 0
            return a / b
        if coef == '自宅療養者等数' and time < format_date1:
            a = self._get_stats(sheet, row, '自宅療養者数', time)
            b = self._get_stats(sheet, row, '社会福士施設等療養者数', time)
            if not isinstance(a, int) or not isinstance(b, int):
                return None
            return a + b

        column = stats_column[coef]
        if column is None:
            return None

        self._validate_header(sheet, time, coef, column, header_row, format_date1)

        v = sheet.cell(row, column).value
        if not isinstance(v, int):
            return None
        return v

    def _read_data_(self, coefs, files):
        data = []
        times = sorted(files,
            key=lambda time_str: [int(x) for x in re.findall('\d+', time_str)])  # 日時の昇順でソート
        for time in times:
            wb = load_workbook(self.data_dir / self._parse_filename(files[time]))
            sheet = wb['公表資料']
            t = datetime.datetime(*[int(x) for x in re.findall('\d+', time)])  # time を datetime 型に
            for row in range(self.data_row_start, self.data_row_end + 1):
                pref = self._parse_pref(sheet[f'{self.pref_column}{row}'])
                for coef in coefs:
                    try:
                        v = self._get_stats(sheet, row, coef, t)
                    except AssertionError as e:
                        print(f'Exception happend. {time}, {row}, {coef}')
                        raise e
                    if v is not None:
                        data.append([pref, t, coef, v])
        df = pd.DataFrame(data, columns=['prefecture', 'time', 'coefficient', 'value'])
        df = df.set_index('time')
        return df

    def _read_data(self, coefs):
        with shelve.open(str(self.data_dir / 'files')) as files:
            df = self._read_data_(coefs, files)
        return df
        
    def show_bed_usage(self, pref, figsize):
        coefs = [
            '(重症者用でない) 確保病床使用率',
            '(重症者用) 確保病床使用率',
            '(宿泊療養施設) 確保居室使用率',
        ]
        df = self._read_data(coefs)
        table = df[df.prefecture == pref][['coefficient', 'value']].pivot(columns='coefficient', values='value')[coefs]
        table.plot(
            figsize=figsize,
            title='ベッドおよび部屋の使用率',
            legend='top_left',
        )

    def show_bed_count(self, pref, figsize):
        coefs = [
            '(重症者用でない) 確保病床数',
            '(重症者用) 確保病床数',
            '(宿泊療養施設) 確保居室数',
        ]
        df = self._read_data(coefs)
        table = df[df.prefecture == pref][['coefficient', 'value']].pivot(columns='coefficient', values='value')[coefs]
        table.plot(
            figsize=figsize,
            title='ベッドおよび部屋の数',
            legend='top_left',
        )

    def show_admission_proportion(self, pref, figsize):
        coefs = [
            '入院率',
            '宿泊療養である割合',
            '自宅療養である割合',
            '療養先調整中である割合',
        ]
        df = self._read_data(coefs)
        table = df[df.prefecture == pref][['coefficient', 'value']].pivot(columns='coefficient', values='value')[coefs]
        table.plot(
            figsize=figsize,
            legend='top_left',
            title='患者の療養等先の割合'
        )

    def show_type_count_1(self, pref, figsize):
        coefs = [
            '療養者数',
            '入院者数',
            '宿泊療養者数',
            '自宅療養者等数',
            '療養先調整中の人数',
        ]
        df = self._read_data(coefs)
        table = df[df.prefecture == pref][['coefficient', 'value']].pivot(columns='coefficient', values='value')[coefs]
        table.plot(
            figsize=figsize,
            legend='top_left',
            logy=True,
            title='療養者数との内訳'
        )

    def show_type_count_2(self, pref, figsize):
        coefs = [
            '療養先調整中の人数',
            '入院先調整中の人数',
        ]
        df = self._read_data(coefs)
        table = df[df.prefecture == pref][['coefficient', 'value']].pivot(columns='coefficient', values='value')[coefs]
        table.plot(
            figsize=figsize,
            legend='top_left',
            logy=True,
        )
